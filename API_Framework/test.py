import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, DoughnutChart, Reference

# ==========================================
# ⚙️ USER SETTINGS
# ==========================================
FILE_PATH = r"D:\Practice_Playwright\API_Framework\Automatic_Sprint_System.xlsx"
CURRENT_SPRINT_NAME = "Sprint-1"
TASKS_TO_GENERATE = 10  # Start with 10, but the sheet is ready for 1000+

# 👥 TEAM LIST (Source for Dropdowns)
employees = [
    "Sagar", "Tushar", "Harsh",
    "Vaibhav", "Shristhi",
    "Deepak", "Aasim",
    "Satyam", "Pushkar",
    "New_Joinee"  # Add names here, they appear in dropdowns automatically
]

# Dropdown Options
priorities = ["High", "Medium", "Low"]
statuses = ["To-Do", "In-Progress", "Pending", "Done"]

# ==========================================
# 🎨 STYLING
# ==========================================
COLOR_BG = "F2F4F8"
COLOR_HEADER = "1F4E78"
COLOR_TEXT = "2D3E50"

font_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="1F4E78", fill_type="solid")

wb = Workbook()
wb.remove(wb.active)

# ==========================================
# 1. PREPARE DATA
# ==========================================
start_date = date.today()
end_date = start_date + timedelta(days=10)
data_rows = []

for t in range(1, TASKS_TO_GENERATE + 1):
    task_id = f"TASK-{100 + t}"
    emp = random.choice(employees)
    desc = f"Feature Validation - Scenario {t}"
    priority = random.choices(priorities, weights=[30, 40, 30], k=1)[0]
    status = random.choices(statuses, weights=[20, 30, 10, 40], k=1)[0]

    if status == "Done":
        pct = 1.0
    elif status == "To-Do":
        pct = 0.0
    else:
        pct = round(random.uniform(0.1, 0.9), 2)

    data_rows.append([task_id, CURRENT_SPRINT_NAME, emp, desc, priority, status, start_date, end_date, pct])

# ==========================================
# 2. SHEET: TEAM SUMMARY (Created FIRST to link Dropdowns)
# ==========================================
ws_summary = wb.create_sheet("Team_Summary")
ws_summary.append(["Employee", "Total", "High", "Done", "Pending"])

# We write employees here so the Task Sheet can "lookup" this list for the dropdown
for i, emp in enumerate(employees, 2):
    ws_summary.append([
        emp,
        f'=COUNTIFS(Task_Assignment!C:C, A{i}, Task_Assignment!B:B, Dashboard!$C$3)',
        f'=COUNTIFS(Task_Assignment!C:C, A{i}, Task_Assignment!B:B, Dashboard!$C$3, Task_Assignment!E:E, "High")',
        f'=COUNTIFS(Task_Assignment!C:C, A{i}, Task_Assignment!B:B, Dashboard!$C$3, Task_Assignment!F:F, "Done")',
        f'=B{i}-D{i}'
    ])

# Totals
lr = len(employees) + 2
ws_summary[f"A{lr}"] = "TOTAL"
ws_summary[f"B{lr}"] = f"=SUM(B2:B{lr - 1})"
ws_summary[f"D{lr}"] = f"=SUM(D2:D{lr - 1})"
ws_summary[f"E{lr}"] = f"=SUM(E2:E{lr - 1})"

# ==========================================
# 3. SHEET: TASK ASSIGNMENT (Fully Automatic)
# ==========================================
ws_tasks = wb.create_sheet("Task_Assignment")
headers = ["Task ID", "Sprint", "Employee", "Description", "Priority", "Status", "Start", "End", "% Done", "Days Left",
           "Flag"]
ws_tasks.append(headers)

# 1. Write Initial Data
for row in data_rows:
    ws_tasks.append(row)
    r = ws_tasks.max_row
    # Formulas
    ws_tasks[f"J{r}"] = f'=IF(F{r}="Done", 0, IF(H{r}="", "", H{r}-TODAY()))'
    ws_tasks[
        f"K{r}"] = f'=IF(AND(J{r}<0, F{r}<>"Done"), "OVERDUE", IF(AND(E{r}="High", F{r}<>"Done"), "ATTENTION", "OK"))'

# 2. CREATE AUTOMATIC DROPDOWNS (Data Validation)

# A. Employee Dropdown (Column C)
# Points to Team_Summary Sheet so it updates if you add people there
# Range: Team_Summary!A2 to A50 (allowing for future hires)
dv_emp = DataValidation(type="list", formula1="Team_Summary!$A$2:$A$50", allow_blank=True)
ws_tasks.add_data_validation(dv_emp)
dv_emp.add("C2:C1000")  # Apply to first 1000 rows

# B. Priority Dropdown (Column E)
dv_pri = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws_tasks.add_data_validation(dv_pri)
dv_pri.add("E2:E1000")  # Apply to first 1000 rows

# C. Status Dropdown (Column F)
dv_stat = DataValidation(type="list", formula1='"To-Do,In-Progress,Pending,Done"', allow_blank=True)
ws_tasks.add_data_validation(dv_stat)
dv_stat.add("F2:F1000")  # Apply to first 1000 rows

# 3. Create Excel Table
# This ensures formulas in J and K automatically appear when you type in a new row
tab = Table(displayName="TaskTable", ref=f"A1:K{ws_tasks.max_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws_tasks.add_table(tab)

# 4. Styling & Column Widths
ws_tasks.column_dimensions["C"].width = 22
ws_tasks.column_dimensions["D"].width = 45
ws_tasks.column_dimensions["F"].width = 15
ws_tasks.column_dimensions["G"].width = 15
ws_tasks.column_dimensions["H"].width = 15

# Conditional Formatting
ws_tasks.conditional_formatting.add(f"F2:F1000", CellIsRule(operator="equal", formula=['"Done"'], stopIfTrue=True,
                                                            fill=PatternFill(start_color="C6EFCE", fill_type="solid")))
ws_tasks.conditional_formatting.add(f"K2:K1000", CellIsRule(operator="equal", formula=['"OVERDUE"'],
                                                            fill=PatternFill(start_color="FFC7CE", fill_type="solid")))

# ==========================================
# 4. SHEET: DASHBOARD
# ==========================================
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_view.showGridLines = False

# Background
for row in ws_dash.iter_rows(min_row=1, max_row=50, max_col=20):
    for cell in row:
        cell.fill = PatternFill(start_color=COLOR_BG, fill_type="solid")

# Header Bar
for row in range(1, 4):
    for col in range(1, 14):
        ws_dash.cell(row=row, column=col).fill = PatternFill(start_color=COLOR_HEADER, fill_type="solid")

ws_dash["B2"] = "AUTOMATIC SPRINT TRACKER"
ws_dash["B2"].font = Font(bold=True, color="FFFFFF", size=18)
ws_dash["J2"] = "Active Sprint:"
ws_dash["J2"].font = Font(color="FFFFFF", bold=True)
ws_dash["K2"] = CURRENT_SPRINT_NAME
ws_dash["K2"].alignment = Alignment(horizontal='center')
ws_dash["K2"].font = Font(bold=True)
ws_dash["K2"].fill = PatternFill(start_color="FFFFFF", fill_type="solid")
ws_dash["C3"] = "=K2"
ws_dash.column_dimensions['C'].hidden = True


# Helper: Card
def create_card(c, r, w, h, bg="FFFFFF"):
    for row_idx in range(r, r + h):
        for col_idx in range(c, c + w):
            ws_dash.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=bg, fill_type="solid")


# KPIs
create_card(2, 5, 3, 2);
ws_dash["B5"] = "TASKS";
ws_dash["B6"] = f"=Team_Summary!B{lr}"
create_card(6, 5, 3, 2);
ws_dash["F5"] = "PROGRESS";
ws_dash["F6"] = f"=IF(Team_Summary!B{lr}=0,0,Team_Summary!D{lr}/Team_Summary!B{lr})"
create_card(10, 5, 3, 2);
ws_dash["J5"] = "PENDING";
ws_dash["J6"] = f"=Team_Summary!E{lr}"

for c in ["B", "F", "J"]:
    ws_dash[f"{c}5"].alignment = Alignment(horizontal="center")
    ws_dash.merge_cells(f"{c}5:{get_column_letter(ws_dash[f'{c}5'].column + 2)}5")
    ws_dash[f"{c}6"].font = Font(size=20, bold=True, color=COLOR_TEXT)
    ws_dash[f"{c}6"].alignment = Alignment(horizontal="center")
    ws_dash.merge_cells(f"{c}6:{get_column_letter(ws_dash[f'{c}6'].column + 2)}6")
ws_dash["F6"].number_format = '0%'
ws_dash["J6"].font = Font(size=20, bold=True, color="D00000")

# Charts
ws_dash["O2"], ws_dash["P2"] = "Done", f"=Team_Summary!D{lr}"
ws_dash["O3"], ws_dash["P3"] = "Pending", f"=Team_Summary!E{lr}"
pie = DoughnutChart();
pie.style = 26;
pie.width = 10;
pie.height = 8
labels = Reference(ws_dash, min_col=15, min_row=2, max_row=3)
data = Reference(ws_dash, min_col=16, min_row=2, max_row=3)
pie.add_data(data, titles_from_data=False);
pie.set_categories(labels)
ws_dash.add_chart(pie, "B9")

bar = BarChart();
bar.type = "col";
bar.style = 4;
bar.height = 8;
bar.width = 20
cats = Reference(ws_summary, min_col=1, min_row=2, max_row=len(employees) + 1)
data = Reference(ws_summary, min_col=2, max_col=4, min_row=1, max_row=len(employees) + 1)
bar.add_data(data, titles_from_data=True);
bar.set_categories(cats)
ws_dash.add_chart(bar, "F9")

# Save
try:
    wb.save(FILE_PATH)
    print("✅ AUTOMATIC SHEET GENERATED!")
    print(f"📂 Location: {FILE_PATH}")
    print("👉 Features:")
    print("   1. Click Cell C12 (or any empty row) -> Employee Dropdown appears.")
    print("   2. Click Cell F12 -> Status Dropdown appears.")
    print("   3. Type a new task -> Formulas for Days Left & Flag auto-generate.")
except PermissionError:
    print("❌ Error: Close the Excel file first.")